"""Build a polished, print-ready Excel tax invoice that mirrors the FSS reference."""

from __future__ import annotations

import os
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from numwords import rupees_in_words
from invoice_core import compute_totals

# --- Reusable styles -------------------------------------------------------
THIN = Side(style="thin", color="404040")
MEDIUM = Side(style="medium", color="1F3864")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NO_BORDER = Border()

HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # deep blue
BANNER_FILL = PatternFill("solid", fgColor="2E5496")   # medium blue
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")    # light blue
LABEL_FILL = PatternFill("solid", fgColor="F2F2F2")    # light grey

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

MONEY_FMT = '#,##0.00'


@dataclass
class InvoiceResult:
    path: str
    subtotal: float
    cgst: float
    sgst: float
    igst: float
    total: float
    supply_type: str


def _set_widths(ws):
    widths = {"A": 9, "B": 54, "C": 18, "D": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _merge(ws, cell_range):
    ws.merge_cells(cell_range)


def generate_invoice(config: dict, client, invoice_number: str,
                     invoice_date: str, line_items: list[dict],
                     output_path: str, document_type: str = "tax") -> InvoiceResult:
    seller = config["seller"]

    t = compute_totals(config, client, line_items)
    subtotal, cgst, sgst, igst = t.subtotal, t.cgst, t.sgst, t.igst
    total, supply_type = t.total, t.supply_type

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    _set_widths(ws)
    ws.sheet_view.showGridLines = False

    r = 1
    # --- Seller header -----------------------------------------------------
    _merge(ws, f"A{r}:D{r}")
    c = ws.cell(r, 1, seller["name"])
    c.font = Font(size=18, bold=True, color="1F3864")
    c.alignment = CENTER
    ws.row_dimensions[r].height = 26
    r += 1

    _merge(ws, f"A{r}:D{r}")
    c = ws.cell(r, 1, f"{seller['address_line1']}, {seller['address_line2']}")
    c.font = Font(size=10)
    c.alignment = CENTER
    r += 1

    _merge(ws, f"A{r}:D{r}")
    c = ws.cell(r, 1,
               f"Mobile: {seller['mobile']}   |   E-mail: {seller['email']}"
               f"   |   GSTIN: {seller['gstin']}")
    c.font = Font(size=10)
    c.alignment = CENTER
    r += 1

    # --- Document banner ---------------------------------------------------
    is_proforma = (document_type or "tax").lower() == "proforma"
    banner_text = "PROFORMA INVOICE" if is_proforma else "TAX INVOICE"
    _merge(ws, f"A{r}:D{r}")
    c = ws.cell(r, 1, banner_text)
    c.font = Font(size=13, bold=True, color="FFFFFF")
    c.alignment = CENTER
    c.fill = BANNER_FILL
    ws.row_dimensions[r].height = 22
    r += 1
    if is_proforma:
        _merge(ws, f"A{r}:D{r}")
        c = ws.cell(r, 1, "This is NOT a Tax Invoice — for approval only. GST will be charged on final tax invoice.")
        c.font = Font(size=9, italic=True, color="B45309")
        c.alignment = CENTER
        ws.row_dimensions[r].height = 18
        r += 1

    # --- Bill-to (left) + Invoice meta (right) -----------------------------
    bill_top = r
    # Left column labels/values
    ws.cell(r, 1, "Bill To:").font = Font(bold=True, size=10)
    ws.cell(r, 3, "Invoice No.:").font = Font(bold=True)
    ws.cell(r, 4, str(invoice_number)).alignment = LEFT
    r += 1

    _merge(ws, f"A{r}:B{r}")
    c = ws.cell(r, 1, client.name)
    c.font = Font(bold=True, size=11)
    c.alignment = LEFT
    ws.cell(r, 3, "Invoice Date:").font = Font(bold=True)
    ws.cell(r, 4, invoice_date).alignment = LEFT
    r += 1

    _merge(ws, f"A{r}:B{r+1}")
    c = ws.cell(r, 1, client.address)
    c.alignment = LEFT_TOP
    c.font = Font(size=10)
    ws.cell(r, 3, "GST Type:").font = Font(bold=True)
    ws.cell(r, 4, "CGST + SGST" if client.mh else "IGST").alignment = LEFT
    r += 1
    ws.cell(r, 3, "Supply Type:").font = Font(bold=True)
    ws.cell(r, 4, "Intra-State" if client.mh else "Inter-State").alignment = LEFT
    r += 1

    _merge(ws, f"A{r}:B{r}")
    c = ws.cell(r, 1, f"GSTIN: {client.gstin}")
    c.font = Font(bold=True, size=10)
    c.alignment = LEFT
    r += 1

    # Box around the bill-to / meta block
    for row in range(bill_top, r):
        for col in range(1, 5):
            ws.cell(row, col).border = Border(
                left=THIN if col == 1 else NO_BORDER.left,
                right=THIN if col == 4 else NO_BORDER.right,
                top=THIN if row == bill_top else NO_BORDER.top,
                bottom=THIN if row == r - 1 else NO_BORDER.bottom,
            )

    r += 1  # spacer

    # --- Line items table --------------------------------------------------
    header_row = r
    headers = ["Sr. No.", "Particulars", "Work Delivered Date", "Amount (Rs.)"]
    for col, text in enumerate(headers, start=1):
        c = ws.cell(header_row, col, text)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER_ALL
    ws.row_dimensions[header_row].height = 22
    r += 1

    for idx, item in enumerate(line_items, start=1):
        ws.cell(r, 1, idx).alignment = CENTER
        pc = ws.cell(r, 2, item["particulars"])
        pc.alignment = LEFT
        ws.cell(r, 3, item.get("date", "")).alignment = CENTER
        amt = ws.cell(r, 4, round(float(item["amount"]), 2))
        amt.alignment = RIGHT
        amt.number_format = MONEY_FMT
        for col in range(1, 5):
            ws.cell(r, col).border = BORDER_ALL
            ws.cell(r, col).font = Font(size=10)
        r += 1

    def _totals_row(label, value, bold=False, fill=None):
        nonlocal r
        _merge(ws, f"A{r}:C{r}")
        lc = ws.cell(r, 1, label)
        lc.alignment = RIGHT
        lc.font = Font(bold=bold, size=10)
        vc = ws.cell(r, 4, round(float(value), 2))
        vc.alignment = RIGHT
        vc.number_format = MONEY_FMT
        vc.font = Font(bold=bold, size=10)
        for col in range(1, 5):
            ws.cell(r, col).border = BORDER_ALL
            if fill:
                ws.cell(r, col).fill = fill
        r += 1

    _totals_row("Sub Total", subtotal)
    for label, amount in t.tax_lines:
        _totals_row(label, amount)
    _totals_row("Grand Total", total, bold=True, fill=TOTAL_FILL)

    # Amount in words
    _merge(ws, f"A{r}:D{r}")
    c = ws.cell(r, 1, "Amount in words: " + rupees_in_words(total))
    c.font = Font(bold=True, italic=True, size=10)
    c.alignment = LEFT
    c.fill = LABEL_FILL
    for col in range(1, 5):
        ws.cell(r, col).border = BORDER_ALL
    ws.row_dimensions[r].height = 20
    r += 2

    # --- Footer: bank details (left) + signature (right) -------------------
    footer_top = r
    bank_lines = [
        "Bank Details:",
        f"Account Name: {seller['name']}",
        f"Bank: {seller['bank_name']}   Branch: {seller['bank_branch']}",
        f"A/C No.: {seller['bank_account']}   IFSC: {seller['bank_ifsc']}",
        f"GSTIN: {seller['gstin']}",
    ]
    for i, line in enumerate(bank_lines):
        _merge(ws, f"A{r}:B{r}")
        c = ws.cell(r, 1, line)
        c.font = Font(bold=(i == 0), size=9)
        c.alignment = LEFT
        r += 1

    # Signature block on the right, aligned with footer top
    sc = ws.cell(footer_top, 3, f"For {seller['name']}")
    sc.font = Font(bold=True, size=10)
    sc.alignment = CENTER
    _merge(ws, f"C{footer_top}:D{footer_top}")
    # Computer-generated note (replaces signature / stamp)
    r = max(r, footer_top + 2) + 1
    _merge(ws, f"A{r}:D{r}")
    note = ws.cell(r, 1,
                  "This is a computer-generated tax invoice and does not "
                  "require a physical signature or company stamp.")
    note.font = Font(italic=True, size=9, color="6B7280")
    note.alignment = CENTER
    ws.row_dimensions[r].height = 18
    r += 1

    # --- Print setup -------------------------------------------------------
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:D{r}"
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    return InvoiceResult(path=output_path, subtotal=subtotal, cgst=cgst,
                        sgst=sgst, igst=igst, total=total,
                        supply_type=supply_type)
