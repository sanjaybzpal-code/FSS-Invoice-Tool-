"""Load the client master list from the live Excel workbook, with a JSON cache fallback."""

from __future__ import annotations

import json
import os
from dataclasses import dataclass, asdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
CACHE_FILE = os.path.join(HERE, "clients_cache.json")
CUSTOM_FILE = os.path.join(HERE, "clients_custom.json")


@dataclass
class Client:
    code: str
    name: str
    gstin: str
    mh: bool
    address: str

    @property
    def mh_label(self) -> str:
        return "YES" if self.mh else "NO"


def _norm(value) -> str:
    return "" if value is None else str(value).strip()


def _parse_mh(value) -> bool:
    return _norm(value).upper().startswith("Y")


def load_from_workbook(workbook_path: str, sheet_name: str) -> list[Client]:
    """Read clients from the CLIENTS sheet.

    Expected columns (row >= 4): Code No. | Client | GST No. | MH YES/NO | Address
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]

    clients: list[Client] = []
    for row in sheet.iter_rows(values_only=True):
        code = _norm(row[0] if len(row) > 0 else "")
        name = _norm(row[1] if len(row) > 1 else "")
        gstin = _norm(row[2] if len(row) > 2 else "")
        mh_raw = row[3] if len(row) > 3 else ""
        address = _norm(row[4] if len(row) > 4 else "")

        # Keep only real client rows: a numeric code and a name.
        if not name or not code or not code[0].isdigit():
            continue
        clients.append(Client(code=code, name=name, gstin=gstin,
                              mh=_parse_mh(mh_raw), address=address))
    wb.close()
    return clients


def save_cache(clients: list[Client]) -> None:
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as fh:
            json.dump([asdict(c) for c in clients], fh, indent=2, ensure_ascii=False)
    except OSError:
        pass


def load_cache() -> list[Client]:
    if not os.path.exists(CACHE_FILE):
        return []
    with open(CACHE_FILE, "r", encoding="utf-8") as fh:
        data = json.load(fh)
    return [Client(**item) for item in data]


def load_custom() -> list[Client]:
    """Manually-added clients saved by the user (kept separate from the workbook)."""
    if not os.path.exists(CUSTOM_FILE):
        return []
    try:
        with open(CUSTOM_FILE, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        return [Client(**item) for item in data]
    except (OSError, json.JSONDecodeError, TypeError):
        return []


def remove_custom_client(name: str) -> None:
    """Remove a manually-added client from the JSON list (by name)."""
    key = (name or "").strip().lower()
    if not key:
        return
    existing = [c for c in load_custom() if c.name.lower() != key]
    with open(CUSTOM_FILE, "w", encoding="utf-8") as fh:
        json.dump([asdict(c) for c in existing], fh, indent=2, ensure_ascii=False)


def rename_custom_client(old_name: str, new_name: str) -> None:
    """Rename a custom client entry if it exists."""
    old_k = (old_name or "").strip().lower()
    new_n = (new_name or "").strip()
    if not old_k or not new_n:
        return
    items = load_custom()
    changed = False
    for c in items:
        if c.name.lower() == old_k:
            c.name = new_n
            changed = True
    if changed:
        with open(CUSTOM_FILE, "w", encoding="utf-8") as fh:
            json.dump([asdict(c) for c in items], fh, indent=2, ensure_ascii=False)


def save_custom_client(name: str, gstin: str, address: str, mh: bool) -> Client:
    """Append a new manually-entered client so it appears in future sessions."""
    existing = load_custom()
    code = f"C{len(existing) + 1}"
    client = Client(code=code, name=name.strip(), gstin=gstin.strip(),
                   mh=bool(mh), address=address.strip())
    # replace if a custom client with the same name already exists
    existing = [c for c in existing if c.name.lower() != client.name.lower()]
    existing.append(client)
    with open(CUSTOM_FILE, "w", encoding="utf-8") as fh:
        json.dump([asdict(c) for c in existing], fh, indent=2, ensure_ascii=False)
    return client


def _merge(workbook_clients: list[Client]) -> list[Client]:
    """Workbook clients first, then custom clients not already present by name."""
    seen = {c.name.lower() for c in workbook_clients}
    merged = list(workbook_clients)
    for c in load_custom():
        if c.name.lower() not in seen:
            merged.append(c)
            seen.add(c.name.lower())
    return merged


def load_clients(workbook_path: str, sheet_name: str) -> tuple[list[Client], str]:
    """Return (clients, source_message).

    Tries the live workbook first; on failure, falls back to the cached copy.
    Custom (manually-added) clients are always merged in.
    """
    try:
        clients = load_from_workbook(workbook_path, sheet_name)
        if clients:
            save_cache(clients)
            merged = _merge(clients)
            extra = len(merged) - len(clients)
            note = f" (+{extra} custom)" if extra else ""
            return merged, f"Loaded {len(clients)} clients from workbook{note}."
        raise ValueError("No client rows found in workbook.")
    except Exception as exc:  # noqa: BLE001 - any read issue should fall back
        cached = load_cache()
        if cached:
            merged = _merge(cached)
            return merged, f"Workbook unavailable ({exc}). Using cached list ({len(merged)})."
        custom = load_custom()
        if custom:
            return custom, f"Workbook unavailable. Using custom clients only ({len(custom)})."
        raise RuntimeError(
            f"Could not read clients from workbook and no cache exists.\n{exc}"
        ) from exc
