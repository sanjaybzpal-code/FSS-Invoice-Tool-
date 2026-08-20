"""PDF and Excel exports for MIS, profitability, and expense reports."""

from __future__ import annotations

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _hdr_fill():
    return PatternFill("solid", fgColor="1F3864")


def _hdr_font():
    return Font(bold=True, color="FFFFFF")


def export_table_excel(path: str, title: str, headers: list[str], rows: list[list]) -> str:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=c)
        cell.fill = _hdr_fill()
        cell.font = _hdr_font()
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def export_segment_pl_excel(path: str, rows: list[dict], seller_name: str = "FSS") -> str:
    headers = ["Segment", "Revenue", "Expense", "Profit", "Profit %"]
    data = []
    for r in rows:
        data.append([
            r.get("BusinessSegmentName", ""),
            float(r.get("Revenue") or 0),
            float(r.get("Expense") or 0),
            float(r.get("Profit") or 0),
            float(r.get("ProfitPercent") or 0),
        ])
    return export_table_excel(path, f"{seller_name} — Segment P&L", headers, data)


def export_expenses_excel(path: str, expenses: list[dict]) -> str:
    headers = ["Date", "Category", "Description", "Amount", "GST", "Total",
               "Vendor", "Segment / Allocation", "Payment Mode", "Reference"]
    rows = []
    for e in expenses:
        rows.append([
            e.get("ExpenseDate", ""),
            e.get("CategoryName", ""),
            e.get("ExpenseDescription", ""),
            float(e.get("Amount") or 0),
            float(e.get("GstAmount") or 0),
            float(e.get("TotalAmount") or 0),
            e.get("VendorName", ""),
            e.get("BusinessSegmentName") or e.get("AllocationType", ""),
            e.get("PaymentMode", ""),
            e.get("ReferenceNumber", ""),
        ])
    return export_table_excel(path, "Expense Register", headers, rows)


def export_mis_excel(path: str, mis: dict) -> str:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "MIS"
    y, m = mis.get("year"), mis.get("month")
    ws.append([f"Monthly MIS — {m:02d}/{y}"])
    ws.append([])
    ws.append(["Segment", "Revenue", "Expense", "Profit", "Profit %", "Outstanding", "Collected"])
    for c in range(1, 8):
        cell = ws.cell(row=3, column=c)
        cell.fill = _hdr_fill()
        cell.font = _hdr_font()
    for r in mis.get("profitability", []):
        seg = next((s for s in mis.get("segments", [])
                    if s["BusinessSegmentId"] == r["BusinessSegmentId"]), {})
        ws.append([
            r.get("BusinessSegmentName"),
            float(r.get("Revenue") or 0),
            float(r.get("Expense") or 0),
            float(r.get("Profit") or 0),
            float(r.get("ProfitPercent") or 0),
            float(seg.get("Outstanding") or 0),
            float(seg.get("TotalCollected") or 0),
        ])
    ws2 = wb.create_sheet("Top Clients")
    ws2.append(["Client", "Revenue"])
    for r in mis.get("top_clients", []):
        ws2.append([r.get("ClientName"), float(r.get("Revenue") or 0)])
    ws3 = wb.create_sheet("Top Expenses")
    ws3.append(["Category", "Amount"])
    for r in mis.get("top_expenses", []):
        ws3.append([r.get("CategoryName"), float(r.get("Expense") or 0)])
    wb.save(path)
    return path


def export_table_pdf(path: str, title: str, headers: list[str], rows: list[list]) -> str:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    doc = SimpleDocTemplate(path, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    data = [headers] + [[str(c) for c in row] for row in rows]
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6fa")]),
    ]))
    story.append(tbl)
    doc.build(story)
    return path


def export_segment_pl_pdf(path: str, rows: list[dict]) -> str:
    headers = ["Segment", "Revenue", "Expense", "Profit", "Profit %"]
    data = []
    for r in rows:
        data.append([
            r.get("BusinessSegmentName", ""),
            f"{float(r.get('Revenue') or 0):,.2f}",
            f"{float(r.get('Expense') or 0):,.2f}",
            f"{float(r.get('Profit') or 0):,.2f}",
            f"{float(r.get('ProfitPercent') or 0):.1f}%",
        ])
    return export_table_pdf(path, "Segment Profitability Report", headers, data)
