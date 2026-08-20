"""Export Client Ledger to PDF and Excel."""

from __future__ import annotations

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

NAVY = colors.HexColor("#1F3864")
THIN = Side(style="thin", color="9aa3b2")


def _money(n) -> str:
    return f"{float(n or 0):,.2f}"


def export_ledger_pdf(client: dict, summary: dict, ledger: list[dict],
                      seller: dict, output_path: str) -> str:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    w = doc.width
    elems = []
    title = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=16,
                           textColor=NAVY, alignment=TA_CENTER)
    sub = ParagraphStyle("s", fontName="Helvetica", fontSize=9, alignment=TA_CENTER)
    elems.append(Paragraph(seller.get("name", "Façade Structural Services"), title))
    elems.append(Paragraph("Client Ledger Statement", sub))
    elems.append(Spacer(1, 8))
    info = (
        f"<b>Client:</b> {client.get('ClientName', '')} &nbsp;|&nbsp; "
        f"<b>GSTIN:</b> {client.get('GSTIN') or '-'} &nbsp;|&nbsp; "
        f"<b>Address:</b> {client.get('Address') or '-'}<br/>"
        f"<b>Tax Invoiced (GST):</b> Rs. {_money(summary.get('TotalTaxInvoiced', 0))} &nbsp;|&nbsp; "
        f"<b>Non-GST Billed:</b> Rs. {_money(summary.get('TotalNonGstBilled', 0))} &nbsp;|&nbsp; "
        f"<b>Cash Received:</b> Rs. {_money(summary.get('TotalReceived', 0))} &nbsp;|&nbsp; "
        f"<b>TDS Deducted:</b> Rs. {_money(summary.get('TotalTdsDeducted', 0))} &nbsp;|&nbsp; "
        f"<b>GST Pending:</b> Rs. {_money(summary.get('GstOutstanding', 0))}<br/>"
        f"<b>Outstanding:</b> Rs. {_money(summary.get('Outstanding', 0))}"
    )
    elems.append(Paragraph(info, sub))
    elems.append(Spacer(1, 10))

    hdr = ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=9,
                         textColor=colors.white, alignment=TA_CENTER)
    data = [[Paragraph("<b>Date</b>", hdr), Paragraph("<b>Voucher No.</b>", hdr),
             Paragraph("<b>Particulars</b>", hdr), Paragraph("<b>Debit</b>", hdr),
             Paragraph("<b>Credit</b>", hdr), Paragraph("<b>Balance</b>", hdr)]]
    cell = ParagraphStyle("c", fontName="Helvetica", fontSize=8)
    cell_r = ParagraphStyle("cr", fontName="Helvetica", fontSize=8, alignment=TA_RIGHT)
    for row in ledger:
        d = row.get("Date")
        ds = d.strftime("%d-%m-%Y") if hasattr(d, "strftime") else str(d or "")
        data.append([
            Paragraph(ds, cell),
            Paragraph(str(row.get("VoucherNo", "")), cell),
            Paragraph(str(row.get("Particulars", "")), cell),
            Paragraph(_money(row.get("Debit", 0)), cell_r),
            Paragraph(_money(row.get("Credit", 0)), cell_r),
            Paragraph(_money(row.get("RunningBalance", 0)), cell_r),
        ])
    col_w = [w * 0.12, w * 0.14, w * 0.40, w * 0.11, w * 0.11, w * 0.12]
    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#9aa3b2")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elems.append(tbl)
    elems.append(Spacer(1, 8))
    elems.append(Paragraph(
        f"<b>Closing Outstanding Balance: Rs. {_money(summary.get('Outstanding', 0))}</b>",
        ParagraphStyle("ob", fontName="Helvetica-Bold", fontSize=10, alignment=TA_RIGHT)))
    elems.append(Spacer(1, 8))
    elems.append(Paragraph(
        "<i>Computer-generated ledger statement — Façade Structural Services</i>",
        ParagraphStyle("f", fontName="Helvetica-Oblique", fontSize=8,
                       textColor=colors.grey, alignment=TA_CENTER)))
    doc.build(elems)
    return output_path


def export_ledger_excel(client: dict, summary: dict, ledger: list[dict],
                        seller: dict, output_path: str) -> str:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"
    ws.merge_cells("A1:F1")
    ws["A1"] = seller.get("name", "Façade Structural Services")
    ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:F2")
    ws["A2"] = "Client Ledger Statement"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A3:F3")
    ws["A3"] = (
        f"Client: {client.get('ClientName','')} | GSTIN: {client.get('GSTIN') or '-'}"
    )
    ws.merge_cells("A4:F4")
    ws["A4"] = (
        f"Tax Invoiced: {_money(summary.get('TotalTaxInvoiced', 0))} | "
        f"Cash: {_money(summary.get('TotalReceived', 0))} | "
        f"TDS: {_money(summary.get('TotalTdsDeducted', 0))} | "
        f"GST Pending: {_money(summary.get('GstOutstanding', 0))} | "
        f"Outstanding: {_money(summary.get('Outstanding', 0))}"
    )
    headers = ["Date", "Voucher No.", "Particulars", "Debit", "Credit", "Running Balance"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(6, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center")
    r = 7
    for row in ledger:
        d = row.get("Date")
        ws.cell(r, 1, d.strftime("%d-%m-%Y") if hasattr(d, "strftime") else d)
        ws.cell(r, 2, row.get("VoucherNo", ""))
        ws.cell(r, 3, row.get("Particulars", ""))
        for col, key in ((4, "Debit"), (5, "Credit"), (6, "RunningBalance")):
            c = ws.cell(r, col, float(row.get(key, 0) or 0))
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right")
        r += 1
    for col, width in zip("ABCDEF", [14, 16, 48, 14, 14, 16]):
        ws.column_dimensions[col].width = width
    wb.save(output_path)
    return output_path
